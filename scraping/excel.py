import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Orijinal CSV dosyasının yolu
original_csv_path = 'C:/Users/bugra/Desktop/turkceldograyan/turkcell_complaints.csv'

# Yeni oluşturulacak Excel dosyasının yolu
output_excel_path = 'C:/Users/bugra/Desktop/turkceldograyan/top_5_categories.xlsx'

# En çok veri olan kategoriler
top_categories = ['Fatura', 'Superbox', 'Hesap', 'İnternet Paketi', 'Yurt Dışı Paketleri']

# Orijinal CSV dosyasını okuma
df = pd.read_csv(original_csv_path, sep=';', encoding='utf-8-sig')

# En çok veri olan kategorilere göre filtreleme
filtered_df = df[df['Category'].isin(top_categories)]

# Yeni Excel dosyasına kaydetme
filtered_df.to_excel(output_excel_path, index=False)

# Excel dosyasını yükleme ve hücre genişliğini ayarlama
wb = load_workbook(output_excel_path)
ws = wb.active

# Hücrelerin genişliğini otomatik olarak ayarlama
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Değişiklikleri kaydetme
wb.save(output_excel_path)

print(f"Filtered data saved to '{output_excel_path}' with adjusted column widths successfully.")
